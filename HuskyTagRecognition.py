from huskylib import HuskyLensLibrary
import pygame
import json
# import matplotlib.pyplot as plt
import time
import os
from openpyxl import Workbook

class Platform(pygame.sprite.Sprite):
    
    def __init__(self, x, y, image):
        super().__init__()
        self.image = image
        self.rect = self.image.get_rect(topleft=(x, y))

# CONSTANTE
WHITE   = (255, 255, 255)
BLUE    = (0, 0, 255)
BLACK   = (0, 0, 0)
ORANGE  = (255, 165, 0)
CONST_PROPORTION    = 4
CONST_ESPACE        = 5

# VARIABLE
quit = False
flagPremierPassage = True
arrayPoint = []
temps = 0
# FIN VARIABLE

# INIT
os.chdir(os.path.dirname(__file__))
pygame.init()
pygame.display.set_caption('Test pygame Berger')
size = width, height = 320 * CONST_PROPORTION, 240 * CONST_PROPORTION
clock = pygame.time.Clock()

#   Lecture du port
hl = HuskyLensLibrary("SERIAL", "COM13")
print(hl.knock())

#   Selection de l'algorithme
hl.algorthim("ALGORITHM_TAG_RECOGNITION")

#   Donne le nombre d'objet enregistré
nbObjet = hl.learnedObjCount()
print(nbObjet)

#   Reçoit tout ce que la caméra voit
obj = hl.requestAll()

#   Création de "Liste"
texts = pygame.sprite.Group()
platforms = pygame.sprite.Group()
suivis = []

#   Création de la police d'écriture du text
idFontBlock = pygame.font.Font('freesansbold.ttf', 10 * CONST_PROPORTION)
txtInit = pygame.font.Font('freesansbold.ttf', 30)

#   Affichage menu console
print("\n\n\n\n")
print("a) Vue globale")
print("b) Déplacement de tag")
# FIN INIT

# MENU
menu = 1
while menu == 1:
    try:
        cmd = input("\nEntrer la lettre de commande:")
        if(cmd == 'a'):
            print("//// Vue globale \\\\\\\\")
            menu = 0
            mode = 1

        if(cmd == 'b'):
            print("//// Déplacement de tag \\\\\\\\")
            print("Liste des tags enregistrés:")
            mode = 2

            count = 1
            obj = hl.requestAll()
            for i in obj:
                if json.dumps(i.learned) == "true":
                    print("\t ID : " + json.dumps(i.ID))
                count+=1

            cmd2 = input("Sélectionner l'ID du tag à suivre:")
            menu = 0

    except KeyboardInterrupt:
        print("\nQUITING")
        quit()
    except IndexError:
        print(f"Commande {v} pas trouvée")
    except Exception as e:
        # General error -> just print it
        print(f"Error {e}")
# FIN MENU

# PYGAME
screen = pygame.display.set_mode(size)
screen.fill(WHITE)
# pygame.display.update()

while not quit:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            quit = True
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_SPACE:
                screen.fill(WHITE)
                arrayPoint.clear()
                flagPremierPassage = True
                temps = 0
                print("\n\r\tReset !")

            if mode == 2:
                if event.key == pygame.K_s:
                    try:
                        wb = Workbook()
                        ws = wb.active
                        ws['A1'] = 'x'
                        ws['B1'] = 'y'
                        ws['C1'] = 't [100 ms]'
                        ws.title = "ValueDeplacement"
                        for pos in arrayPoint:
                            ws.append(pos)
                        wb.save("PositionXYZ.xlsx")
                        print("\n\r\tSauvegardé !")
                    except PermissionError:
                        print("\n\r\tERROR: Le fichier est en cours de lecture !")

    texts.empty()
    platforms.empty()
    if(mode == 1):
        screen.fill(WHITE)
        obj = hl.requestAll()
        for i in obj:
            block_tag = pygame.Surface((int(json.dumps(i.width)) * CONST_PROPORTION, int(json.dumps(i.height)) * CONST_PROPORTION))
            block_tag.fill(pygame.Color('black'))

            if json.dumps(i.learned) == "true":
                text = idFontBlock.render('ID:{}'.format(int(json.dumps(i.ID))), True, (255, 255, 255), (0, 0, 0))

            else:
                text = idFontBlock.render('-', True, (255, 0, 0), (0, 0, 0))

            texts.add(Platform(int(json.dumps(i.x)) * CONST_PROPORTION, int(json.dumps(i.y)) * CONST_PROPORTION, text))
            platforms.add(Platform(int(json.dumps(i.x)) * CONST_PROPORTION, int(json.dumps(i.y)) * CONST_PROPORTION, block_tag))

    if(mode == 2):
        textInit = txtInit.render('Save = s     Reset = SPACE', True, (0, 0, 0), (255, 255, 255))
        screen.blit(textInit, (0, 0))
        obj = hl.requestAll()
        for i in obj:
            if(int(json.dumps(i.ID)) == int(cmd2)):
                # block_tag = pygame.Surface((int(json.dumps(i.width)) * CONST_PROPORTION, int(json.dumps(i.height)) * CONST_PROPORTION))
                # block_tag.fill(pygame.Color('black'))
                pos = (int(json.dumps(i.x)) * CONST_PROPORTION, int(json.dumps(i.y)) * CONST_PROPORTION, temps)
                if not flagPremierPassage:
                    # if pos not in arrayPoint:
                        if (pos[0] > (arrayPoint[-1][0] + CONST_ESPACE)) or (pos[0] < (arrayPoint[-1][0] - CONST_ESPACE)):
                            if (pos[1] > (arrayPoint[-1][1] + CONST_ESPACE)) or (pos[1] < (arrayPoint[-1][1] - CONST_ESPACE)):
                                arrayPoint.append(pos)
                                pygame.draw.circle(screen, BLUE, (pos[0], pos[1]), 2)
                                start_pos   = (arrayPoint[-2][0], arrayPoint[-2][1])
                                end_pos     = (arrayPoint[-1][0], arrayPoint[-1][1])
                                pygame.draw.line(screen, ORANGE, start_pos, end_pos, width=2)
                else:
                    flagPremierPassage = False
                    arrayPoint.append(pos)
                    pygame.draw.circle(screen, BLUE, (pos[0], pos[1]), 2)

                # print(arrayPoint)
                time.sleep(0.1) # Wait 100ms
                temps += 1

                if json.dumps(i.learned) == "true":
                    # text = idFontBlock.render('ID:{}'.format(int(json.dumps(i.ID))), True, (255, 255, 255), (0, 0, 0))
                    text = idFontBlock.render('ID:{}'.format(int(json.dumps(i.ID))), True, (255, 255, 255), (0, 0, 0))

                else:
                    text = idFontBlock.render('-', True, (255, 0, 0), (0, 0, 0))

                # texts.add(Platform(int(json.dumps(i.x)) * CONST_PROPORTION, int(json.dumps(i.y)) * CONST_PROPORTION, text))
                # platforms.add(Platform(int(json.dumps(i.x)) * CONST_PROPORTION, int(json.dumps(i.y)) * CONST_PROPORTION, block_tag))
                
    
    platforms.update()
    texts.update()
        
    platforms.draw(screen)
    texts.draw(screen)

    pygame.display.flip()
    pygame.display.update()
    clock.tick(30)

pygame.quit()
quit()
